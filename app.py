from flask import Flask, render_template, request, session, redirect, url_for
import openpyxl
import os
from datetime import datetime


# ------------------- Flask Setup -------------------
app = Flask(__name__)
app.secret_key = 'your_secret_key'  # change to a secure key

EXCEL_FILE = 'data.xlsx'  # Excel file path
sheet_name='users'

# ------------------- Initialize Excel Workbook -------------------
if not os.path.exists(EXCEL_FILE):
    wb = openpyxl.Workbook()
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']

    # Create initial sheets
    users_sheet = wb.create_sheet('users')
    users_sheet.append(['username', 'email', 'password'])

    feedback_sheet = wb.create_sheet('feedback')
    feedback_sheet.append(['username', 'feedback'])

    wb.save(EXCEL_FILE)
    print("✅ Excel file created with initial sheets.")
else:
    print("📄 Excel file already exists.")


# ------------------- Excel Helpers -------------------
def load_data(sheet_name):
    """Load data from a sheet, return list of rows (excluding headers)."""
    if os.path.exists(EXCEL_FILE):
        wb = openpyxl.load_workbook(EXCEL_FILE)
        if sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            return [list(row) for row in sheet.iter_rows(min_row=2, values_only=True)]
    return []

def save_data(data, sheet_name, headers):
    """Save data to a sheet safely, preserving other sheets."""
    if os.path.exists(EXCEL_FILE):
        wb = openpyxl.load_workbook(EXCEL_FILE)
    else:
        wb = openpyxl.Workbook()
        if 'Sheet' in wb.sheetnames:
            del wb['Sheet']

    if sheet_name in wb.sheetnames:
        del wb[sheet_name]

    ws = wb.create_sheet(sheet_name)
    
    # Append headers if provided and not empty
    if headers and all(h is not None for h in headers):
        ws.append(list(headers))

    # Append all data rows
    for row in data:
        ws.append(list(row))

    wb.save(EXCEL_FILE)

# ------------------- Routes -------------------

# Home Page
@app.route('/home')
def home():
        return render_template('home.html')


# Dashboard Page
@app.route('/dashboard')
def dashboard():
        return render_template('dashboard.html')

# Dashboard Page
@app.route('/apptitude')
def apptitude():
        return render_template('apptitude.html')

# Dashboard Page
@app.route('/GD')
def GD():
        return render_template('GD.html')

# Dashboard Page
@app.route('/teqchnical')
def teqchnical():
        return render_template('teqchnical.html')

# Dashboard Page
@app.route('/HR')
def HR():
        return render_template('HR.html')

# Dashboard Page
@app.route('/Resume')
def Resume():
        return render_template('Resume.html')


@app.route('/')
def index():
    return redirect('/signup')

# Signup Page
@app.route('/signup', methods=['GET', 'POST'])
def signup():
    if request.method == 'POST':
        username = request.form['username']
        email = request.form['email']
        password = request.form['password']

        users_data = load_data('users')
        # Check if email already exists
        for row in users_data:
            if row[1] == email:
                return render_template('signup.html', error='Email already registered')

        new_user = [username, email, password]
        updated_users = users_data + [new_user]
        save_data(updated_users, sheet_name='users', headers=['username', 'email', 'password'])

        session['username'] = username  # Auto login after signup
        return redirect(url_for('home'))

    return render_template('signup.html')

# Login Page
@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']

        users_data = load_data('users')
        for row in users_data:
            if row[0] == username and row[2] == password:
                session['username'] = username
                return redirect(url_for('home'))
        return render_template('login.html', error='Invalid credentials')
    return render_template('login.html')

# Logout
@app.route('/logout')
def logout():
    session.pop('username', None)
    return redirect(url_for('login'))

# Feedback Page
@app.route('/feedback', methods=['GET', 'POST'])
def feedback():
    if request.method == 'POST':
        name = request.form.get('name')
        email = request.form.get('email')
        message = request.form.get('message')
        timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

        file_path = 'data.xlsx'

        # Create file if it doesn't exist
        if not os.path.exists(file_path):
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append(['Name', 'Email', 'Message', 'Timestamp'])
            wb.save(file_path)

        # Append feedback
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        ws.append([name, email, message, timestamp])
        wb.save(file_path)

        # Redirect back to feedback page
        return redirect('/feedback')

    return render_template('feedback.html')


"""
@app.route('/feedback', methods=['GET', 'POST'])
def feedback():
    if 'username' not in session:
        return redirect(url_for('login'))

    if request.method == 'POST':
        feedback_text = request.form['feedback']
        username = session['username']

        feedback_data = load_data('feedback')
        new_feedback = [username, feedback_text]
        updated_feedback = feedback_data + [new_feedback]
        save_data(updated_feedback, sheet_name='feedback', headers=['username', 'feedback'])

        return redirect(url_for('dashboard'))
    return render_template('feedback.html')"""

# Placements Page
@app.route('/placement')
def placement():
  
        return render_template('placement.html')
   

# Internships Page
@app.route('/internship')
def internship():
        return render_template('internship.html')
    

# Branches Page
@app.route('/branch')
def branch():
    
        return render_template('branch.html')


# Alumni Page
@app.route('/alumini')
def alumini():
        return render_template('alumini.html')
    

# TPO Page
@app.route('/TPO')
def TPO():
    
        return render_template('TPO.html')
   
@app.route('/<page>.html')
def html_pages(page):
    try:
        return render_template(f'{page}.html')
    except:
        return "Page not found", 404

# ------------------- Run App -------------------
if __name__ == '__main__':
    app.run(debug=True, port=5000)
